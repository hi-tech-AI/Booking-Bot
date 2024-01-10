from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.chrome.options import Options
from time import sleep
from openpyxl import Workbook

options = Options()
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.130 Safari/537.3")
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
service = Service(executable_path = "C:\chromedriver-win64\chromedriver.exe")   
driver = webdriver.Chrome(options = options)

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

# driver.get('https://icp.administracionelectronica.gob.es/icpplus/index.html')

wb = Workbook()
sheet = wb.active
item = ['No', 'Nationality']

for i in range(0, 2):
    sheet.cell(row = 1, column = i + 1).value = item[i]

nationalitys = Find_Elements(driver, By.TAG_NAME, 'option')

print(len(nationalitys))

start_row = 2
for nationality in nationalitys:
    sheet.cell(row = start_row, column = 1).value = start_row - 1
    sheet.cell(row = start_row, column = 2).value = nationality.text
    start_row += 1

wb.save('nationality.xlsx')